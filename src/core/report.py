"""xlsx report generator — three-sheet evaluation report.

Sheet 1 · Overview: pass rate, per-layer averages, attribution histogram
Sheet 2 · Per-case detail: color-coded by attribution level
Sheet 3 · Blocker analysis: root cause + suggested fix per blocker case
"""

from __future__ import annotations

import logging
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from src.core.models import EvalResult, FailureLevel

logger = logging.getLogger(__name__)

ATTRIBUTION_COLORS = {
    FailureLevel.INTERFACE_EXCEPTION: "FF444444",
    FailureLevel.LLM_DECOMPOSITION_ERROR: "FFE74C3C",
    FailureLevel.COLOR_LOSS: "FFFF8C42",
    FailureLevel.PERCENTAGE_LOSS: "FFFF8C42",
    FailureLevel.FIELD_MISSING: "FF9B59B6",
    FailureLevel.SCENE_MISSING: "FF9B59B6",
    FailureLevel.CORRECT_BUT_NO_DATA: "FFF1C40F",
    FailureLevel.CORRECT: "FF27AE60",
}

LAYER_NAMES = ["L1", "L2", "L3", "L4", "L5", "L6"]


def generate_report(
    results: list[EvalResult],
    output_path: str | Path,
    *,
    case_map: dict[str, str] | None = None,
) -> Path:
    """Generate a 3-sheet xlsx report from evaluation results.

    Parameters
    ----------
    results : list[EvalResult]
        Evaluation results from SixLayerEvaluator.run().
    output_path : str | Path
        Where to write the xlsx file.
    case_map : dict[str, str] | None
        Optional mapping from case_id to human-readable query text.

    Returns
    -------
    Path
        The path to the generated xlsx file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise ImportError(
            "openpyxl is required for report generation. "
            "Install with: pip install openpyxl"
        ) from e

    wb = Workbook()
    case_map = case_map or {}

    _build_overview_sheet(wb.active, results)
    _build_detail_sheet(wb.create_sheet("Case Detail"), results, case_map)
    _build_blocker_sheet(wb.create_sheet("Blocker Analysis"), results)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    logger.info("Report saved to %s", output_path)
    return output_path


def _build_overview_sheet(ws: Any, results: list[EvalResult]) -> None:
    """Sheet 1: Overview — pass rate, averages, attribution histogram."""
    from openpyxl.styles import Alignment, Font, PatternFill

    ws.title = "Overview"

    header_fill = Font(name="Calibri", bold=True, size=14, color="FF1A1A2E")
    ws["A1"] = "AgentLens Evaluation Report"
    ws["A1"].font = header_fill
    ws.merge_cells("A1:D1")

    ws["A3"] = "Generated"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A4"] = "Total Cases"
    ws["B4"] = len(results)

    passed = sum(1 for r in results if r.overall_passed)
    blocked = sum(1 for r in results if r.blocker is not None)
    ws["A5"] = "Pass Rate"
    ws["B5"] = f"{passed}/{len(results)} ({passed / len(results) * 100:.1f}%)" if results else "N/A"
    ws["A6"] = "Blocker Cases"
    ws["B6"] = blocked

    ws["A9"] = "Per-Layer Average Scores"
    ws["A9"].font = Font(bold=True)
    for i, layer in enumerate(LAYER_NAMES):
        scores = [r.layer_scores[layer].score for r in results if layer in r.layer_scores]
        avg = sum(scores) / len(scores) if scores else 0.0
        ws.cell(row=10 + i, column=1, value=layer)
        ws.cell(row=10 + i, column=2, value=round(avg, 3))

    ws["A17"] = "Attribution Distribution"
    ws["A17"].font = Font(bold=True)
    attr_counter: Counter[str] = Counter()
    for r in results:
        attr_counter[r.attribution_level] += 1

    for i, (level, count) in enumerate(sorted(attr_counter.items(), key=lambda x: -x[1])):
        ws.cell(row=18 + i, column=1, value=level)
        ws.cell(row=18 + i, column=2, value=count)
        color = ATTRIBUTION_COLORS.get(
            FailureLevel(level), "FF808080"
        )
        cell = ws.cell(row=18 + i, column=3, value="■")
        cell.font = Font(color=color, size=16)

    avg_latency = sum(r.latency_ms for r in results) / len(results) if results else 0
    ws["A27"] = "Avg Latency (ms)"
    ws["B27"] = round(avg_latency, 1)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 8


def _build_detail_sheet(ws: Any, results: list[EvalResult], case_map: dict[str, str]) -> None:
    """Sheet 2: Per-case detail, color-coded by attribution."""
    from openpyxl.styles import Alignment, Font, PatternFill

    headers = [
        "Case ID", "Query", "L1", "L2", "L3", "L4", "L5", "L6",
        "Attribution", "Blocker", "Latency(ms)",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill(start_color="FF1A1A2E", end_color="FF1A1A2E", fill_type="solid")

    for row_idx, result in enumerate(sorted(results, key=lambda r: r.case_id), 2):
        ws.cell(row=row_idx, column=1, value=result.case_id)
        ws.cell(row=row_idx, column=2, value=case_map.get(result.case_id, "")[:60])

        for col_idx, layer in enumerate(LAYER_NAMES, 3):
            score = result.layer_scores.get(layer)
            if score:
                cell = ws.cell(row=row_idx, column=col_idx, value=round(score.score, 2))
                if score.score >= 0.8:
                    cell.fill = PatternFill(start_color="FFD5F5E3", end_color="FFD5F5E3", fill_type="solid")
                elif score.score >= 0.5:
                    cell.fill = PatternFill(start_color="FFFEF9E7", end_color="FFFEF9E7", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFFADBD8", end_color="FFFADBD8", fill_type="solid")
            else:
                ws.cell(row=row_idx, column=col_idx, value="-")

        ws.cell(row=row_idx, column=9, value=result.attribution_level)
        color = ATTRIBUTION_COLORS.get(
            FailureLevel(result.attribution_level) if result.attribution_level != "unknown" else FailureLevel.CORRECT,
            "FF808080",
        )
        attr_cell = ws.cell(row=row_idx, column=9)
        attr_cell.font = Font(color=color)

        ws.cell(row=row_idx, column=10, value=result.blocker or "")
        ws.cell(row=row_idx, column=11, value=round(result.latency_ms, 1))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_col_letter_ws(col)].width = [18, 40, 7, 7, 7, 7, 7, 7, 22, 8, 12][col - 1]


def _build_blocker_sheet(ws: Any, results: list[EvalResult]) -> None:
    """Sheet 3: Blocker analysis — root cause + fix for each blocker case."""
    from openpyxl.styles import Font, PatternFill

    headers = ["Blocker Case", "Blocker Layer", "Root Cause", "Suggested Fix"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill(start_color="FF1A1A2E", end_color="FF1A1A2E", fill_type="solid")

    blockers = [r for r in results if r.blocker is not None]
    if not blockers:
        ws.cell(row=2, column=1, value="No blocker cases — all passed!")
        ws.cell(row=2, column=1).font = Font(color="FF27AE60", bold=True)
    else:
        for row_idx, result in enumerate(blockers, 2):
            ws.cell(row=row_idx, column=1, value=result.case_id)
            ws.cell(row=row_idx, column=2, value=result.blocker)
            root_cause = result.attribution.root_cause if result.attribution else "Unknown"
            fix = result.attribution.suggested_fix if result.attribution else "N/A"
            ws.cell(row=row_idx, column=3, value=root_cause)
            ws.cell(row=row_idx, column=4, value=fix)

            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color="FFFADBD8", end_color="FFFADBD8", fill_type="solid"
                )

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 55


def get_col_letter_ws(col: int) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(col)